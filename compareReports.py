import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import math
import os
import warnings
warnings.simplefilter(action="ignore", category=FutureWarning)

# ───────────────────────── COLOR PALETTE ──────────────────────────────
COLOR_HEADER = "8DB4E2"  # header background
COLOR_HEADER_FONT = "000000"  # header text color
COLOR_SECTION = "DAEEF3"  # parent/section row
COLOR_ZEBRA = "F7F7F7"  # zebra separator
COLOR_HIGHLIGHT = "FFC000"  # highlight differences

MAX_COL_WIDTH = 120  # maximum column width when adjusting
PADDING = 2  # column padding


# --------------------- Adjust Columns -----------------------------------------
def fit_columns(ws, max_width: int = MAX_COL_WIDTH, padding: int = PADDING):
    for idx, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
        max_len = 0
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        width = min(max_len + padding, max_width)
        letter = get_column_letter(idx)
        ws.column_dimensions[letter].width = width
        if width == max_width:
            for cell in col:
                cell.alignment = Alignment(wrap_text=True)



def combine_and_format(
        input1: str,
        input2: str,
        output: str,
        error_diff_thresholds,
        avg_diff_thresholds,
        percentile_diff_thresholds,
        sheet: str = "metrics"

):
    # --- 2.1) Read and tag columns ---
    df1 = pd.read_excel(input1, sheet_name=sheet, header=1)
    df2 = pd.read_excel(input2, sheet_name=sheet, header=1)
    keep_columns = ["Label", "#Samples", "FAIL", "Error %", "Average", "90th pct"]
    df1 = df1[keep_columns]
    df2 = df2[keep_columns]


    df1_pref = df1.rename(columns={c: f"{input1.split('.')[0]}\n{c}" for c in df1.columns if c != "Label"})
    df2_pref = df2.rename(columns={c: f"{input2.split('.')[0]}\n{c}" for c in df2.columns if c != "Label"})

    # --- 2.2) Find common, only in 1 and only in 2 ---
    s1 = set(df1["Label"].dropna().astype(str).str.strip())
    s2 = set(df2["Label"].dropna().astype(str).str.strip())
    only1, only2 = s1 - s2, s2 - s1

    df_common = pd.merge(df1_pref, df2_pref, on="Label", how="inner")
    df_o1 = df1_pref[df1_pref["Label"].isin(only1)].copy()
    df_o2 = df2_pref[df2_pref["Label"].isin(only2)].copy()

    # complete columns that may be missing after the merge
    for c in df2_pref.columns:
        if c != "Label" and c not in df_o1:
            df_o1[c] = pd.NA
    for c in df1_pref.columns:
        if c != "Label" and c not in df_o2:
            df_o2[c] = pd.NA

    combined = (
        pd.concat([df_common, df_o1, df_o2], ignore_index=True)
        .dropna(how="all")
        .sort_values(by="Label", key=lambda s: s.str.lower())
        .reset_index(drop=True)
    )

    # --- 2.3) Re‑order columns ---
    base = [c for c in df1.columns if c != "Label"]
    cols = ["Label"]
    for m in base:
        cols += [f"{input1.split('.')[0]}\n{m}", f"{input2.split('.')[0]}\n{m}"]
    combined = combined[cols]

    # ── 2.4) Round numeric columns to 2 decimals ─────────────────────────────
    numeric_cols = combined.select_dtypes(include="number").columns

    def fmt_num(x):
        if pd.isna(x):
            return x
        if isinstance(x, (int, float)) and float(x).is_integer():
            return int(x)
        return round(x, 2)

    combined[numeric_cols] = combined[numeric_cols].applymap(fmt_num)

    # --- 2.5) Export to Excel ---
    combined.to_excel(output, sheet_name=sheet, index=False)

    # --- 2.6) openpyxl formatting ---
    wb = load_workbook(output)
    ws = wb[sheet]

    # basic styles
    header_fill = PatternFill("solid", fgColor=COLOR_HEADER)
    thin = Side("thin", color="000000")
    border = Border(thin, thin, thin, thin)

    # Headers with white font
    for col_idx in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        cell.fill = header_fill
        cell.font = Font(color=COLOR_HEADER_FONT, bold=True)
        cell.border = border

    # Borders and two‑decimal number format in numeric columns
    num_format_cols = {
        col_idx
        for col_idx, col_name in enumerate(cols, start=1)
        if col_name in numeric_cols
    }
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(cols)):
        for cell in row:
            cell.border = border
            if cell.column in num_format_cols and isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"

    # 7.2) Highlight significant differences in amber
    highlight_fill = PatternFill("solid", fgColor=COLOR_HIGHLIGHT)
    col_map = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    error1_col = col_map.get(f"{input1.split('.')[0]}\nError %")
    error2_col = col_map.get(f"{input2.split('.')[0]}\nError %")
    avg1_col = col_map.get(f"{input1.split('.')[0]}\nAverage")
    avg2_col = col_map.get(f"{input2.split('.')[0]}\nAverage")
    p90th1_col = col_map.get(f"{input1.split('.')[0]}\n90th pct")
    p90th2_col = col_map.get(f"{input2.split('.')[0]}\n90th pct")

    for row in range(2, ws.max_row + 1):
        try:
            if error1_col and error2_col:
                v1 = ws.cell(row=row, column=error1_col).value
                v2 = ws.cell(row=row, column=error2_col).value
                if v1 is not None and v2 is not None and (float(v1) - float(v2)) > error_diff_thresholds:
                    ws.cell(row=row, column=error1_col).fill = highlight_fill
                    ws.cell(row=row, column=error2_col).fill = highlight_fill
                if v1 is not None and v2 is not None and (float(v1) - float(v2)) < -error_diff_thresholds:
                    ws.cell(row=row, column=error1_col).fill = PatternFill(fill_type="solid", fgColor="90EE90")
                    ws.cell(row=row, column=error2_col).fill = PatternFill(fill_type="solid", fgColor="90EE90")

            if avg1_col and avg2_col:
                a1 = ws.cell(row=row, column=avg1_col).value
                a2 = ws.cell(row=row, column=avg2_col).value
                if a1 is not None and a2 is not None and (float(a1) - float(a2)) > avg_diff_thresholds:
                    ws.cell(row=row, column=avg1_col).fill = highlight_fill
                    ws.cell(row=row, column=avg2_col).fill = highlight_fill
                if a1 is not None and a2 is not None and (float(a1) - float(a2)) < -avg_diff_thresholds:
                    ws.cell(row=row, column=avg1_col).fill = PatternFill(fill_type="solid", fgColor="90EE90")
                    ws.cell(row=row, column=avg2_col).fill = PatternFill(fill_type="solid", fgColor="90EE90")

            if p90th1_col and p90th2_col:
                v1 = ws.cell(row=row, column=p90th1_col).value
                v2 = ws.cell(row=row, column=p90th2_col).value
                if v1 is not None and v2 is not None and (float(v1) - float(v2)) > percentile_diff_thresholds:
                    ws.cell(row=row, column=p90th1_col).fill = highlight_fill
                    ws.cell(row=row, column=p90th2_col).fill = highlight_fill
                if v1 is not None and v2 is not None and (float(v1) - float(v2)) < -percentile_diff_thresholds:
                    ws.cell(row=row, column=p90th1_col).fill = PatternFill(fill_type="solid", fgColor="90EE90")
                    ws.cell(row=row, column=p90th2_col).fill = PatternFill(fill_type="solid", fgColor="90EE90")

        except Exception as e:
            print(f"⚠️ Error comparing row {row}: {e}")

    # 7.3) Insert separator rows (zebra)
    zebra_fill = PatternFill("solid", fgColor=COLOR_ZEBRA)
    group_rows = []
    last_prefix = None
    for r in range(2, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if val is None or not str(val).strip():
            continue
        pref = str(val).split(".")[0]
        if last_prefix is not None and pref != last_prefix:
            group_rows.append(r)
        last_prefix = pref
    for idx in sorted(group_rows, reverse=True):
        ws.insert_rows(idx)
        for c in range(1, len(cols) + 1):
            ws.cell(idx, c).fill = zebra_fill

    # 7.4) Automatic grouping (parent rows)
    summary_fill = PatternFill("solid", fgColor=COLOR_SECTION)
    current_summary = None
    for r in range(2, ws.max_row + 1):
        lbl = ws.cell(r, 1).value
        if lbl is None:
            current_summary = None
            continue
        if "-" not in str(lbl):
            current_summary = r
            for cell in ws[r]:
                cell.fill = summary_fill
        else:
            if current_summary:
                ws.row_dimensions[r].outlineLevel = 1
                ws.row_dimensions[r].hidden = True
                ws.row_dimensions[r].collapsed = True

    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.freeze_panes = "A2"

    # 7.5) Auto‑fit columns and save
    fit_columns(ws)
    wb.save(output)
    print(f"✅ {output} created and formatted.")


# ───────────────────────── 3) MAIN ────────────────────────────────────────────


def compareReports():
    while True:
        file1 = input('Insert the first report for comparison: ')
        if not os.path.exists(file1):
            print("The file does not exist. Make sure you include the file extension.")
            continue
        break
    while True:
        file2 = input('Insert the second report for comparison: ')
        if not os.path.exists(file2):
            print("The file does not exist. Make sure you include the file extension.")
            continue
        break
    while True:
        try:
            error_diff_thresholds = float(input('Max allowed difference in Error Rate (%): ').strip())
        except:
            print("Only float values are allowed.")
            continue
        break
    while True:
        try:
            avg_diff_thresholds = int(input('Max allowed difference in Avg Response Time (ms): ').strip())
        except:
            print("Only integer values are allowed.")
            continue
        break
    while True:
        try:
            percentile_diff_thresholds = int(input('Max allowed difference in 90th Percentile (ms): ').strip())
        except:
            print("Only integer values are allowed.")
            continue
        break

    combine_and_format(
        input1=file1,
        input2=file2,
        output=f"{file1.split('.')[0]}_vrs_{file2.split('.')[0]}.xlsx",
        sheet="metrics",
        error_diff_thresholds=error_diff_thresholds,
        avg_diff_thresholds=avg_diff_thresholds,
        percentile_diff_thresholds=percentile_diff_thresholds
    )


if __name__ == "__main__":
    compareReports()
