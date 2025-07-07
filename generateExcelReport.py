
import warnings
import requests
import json
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from bs4 import BeautifulSoup  # To parse HTML
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from time import sleep
from colorama import Fore, Style



MAX_COL_WIDTH = 120  # Approx. pixels (tweak as needed)
PADDING = 2  # Extra space

def fit_columns(ws, max_width: int = MAX_COL_WIDTH, padding: int = PADDING):
    """
    Resize every column in a worksheet, enforcing a maximum width and
    enabling wrap-text when the limit is reached.
    """
    for idx, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
        max_len = 0
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        width = min(max_len + padding, max_width)
        col_letter = get_column_letter(idx)
        ws.column_dimensions[col_letter].width = width

        # If the width hit the limit, wrap the cell contents
        if width == max_width:
            for cell in col:
                cell.alignment = Alignment(wrap_text=True)

def extract_json(text: str, start_index: int):
    """
    Starting from a given index, find the first '{' and walk forward,
    counting braces to return the entire JSON block—even if strings
    contain braces.
    """
    json_start = text.find('{', start_index)
    if json_start == -1:
        return None
    count = 0
    index = json_start
    in_string = False
    escape = False
    while index < len(text):
        char = text[index]
        if char == '"' and not escape:
            in_string = not in_string
        if not in_string:
            if char == '{':
                count += 1
            elif char == '}':
                count -= 1
                if count == 0:
                    return text[json_start:index + 1]
        escape = (char == '\\' and not escape)
        index += 1
    return None
# ============================================================
# 0. Extract Start Time and End Time from the summary page
# ============================================================
def genarateExcelreport():
    while True:
        report_url = input(Fore.RESET+"Insert the JMeter Report URL: ")
        try:
            report_response = requests.get(report_url)
            if report_response.status_code != 200:
                print("Error retrieving the summary page, status code:", report_response.status_code)
                exit(1)
            html_text = report_response.text
            break
        except Exception:
            print(Fore.RED + "❌ Invalid URL. Please enter a valid JMeter HTML report (ends with /index.html).")


    # Ask the user whether SLA highlighting is required
    while True:
        highlight_sla_flag = input("Do you want to highlight values above SLA? (y/n): ").lower()
        if highlight_sla_flag in ('y', 'n'):
            break
        print("Answer with 'y' for yes or 'n' for no.")

    if highlight_sla_flag == 'y':
        while True:
            try:
                avg_time_sla = float(input("Insert Average Response Time SLA (ms): "))
                break
            except ValueError:
                print("A numeric value is expected.")

        while True:
            try:
                error_sla = float(input("Insert the Error Rate SLA (%): "))
                break
            except ValueError:
                print("A numeric value is expected.")

        while True:
            include_parent_transactions = input(
                "Apply SLA values to parent transactions as well? (default applies only to children) y/n: "
            ).lower()
            if include_parent_transactions == 'y':
                include_parent_transactions = True
                break
            elif include_parent_transactions == 'n':
                include_parent_transactions = False
                break
            print("Answer with 'y' for yes or 'n' for no.")
    else:
        include_parent_transactions = False
        error_sla = float("inf")
        avg_time_sla = float("inf")

    if highlight_sla_flag=='y':
        print(Fore.RESET+f"""Generating Excel report with the following configuration:     
        Report URL: {report_url}
        Avg Response SLA: {avg_time_sla} ms
        Error Rate SLA: {error_sla}%
        Apply SLA to parents: {include_parent_transactions}\n\n""")
    else:
        print(Fore.RESET + f"""Generating Excel report with the following configuration:
        Report URL: {report_url}
        Avg Response SLA: not applied
        Error Rate SLA: not applied
        Apply SLA to parents: not applied\n\n""")

    # Parse general information
    soup = BeautifulSoup(html_text, "html.parser")
    general_table = soup.find("table", id="generalInfos")
    start_time = ""
    end_time = ""

    if general_table:
        for row in general_table.find_all("tr"):
            cells = row.find_all("td")
            if len(cells) >= 2:
                label = cells[0].get_text(strip=True)
                value = cells[1].get_text(strip=True).strip('"')
                if label == "Start Time":
                    start_time = value
                elif label == "End Time":
                    end_time = value
    else:
        print("Table 'generalInfos' was not found on the summary page.")

    # ============================================================
    # 1. Download dashboard.js and extract the statisticsTable JSON
    # ============================================================
    js_url = report_url.split('index')[0] + "content/js/dashboard.js"
    js_response = requests.get(js_url)
    if js_response.status_code != 200:
        print("The URL must point to the dashboard view. Status code:", js_response.status_code)
        sleep(5)
        exit(1)
    js_content = js_response.text

    search_string = 'createTable($("#statisticsTable")'
    position = js_content.find(search_string)
    if position == -1:
        print("Could not find createTable call for statisticsTable.")
        exit(1)
    comma_position = js_content.find(',', position)
    if comma_position == -1:
        print("Could not find the comma after the selector for statisticsTable.")
        exit(1)

    json_text = extract_json(js_content, comma_position)
    if not json_text:
        print("Failed to extract the JSON block for statisticsTable.")
        exit(1)

    try:
        data = json.loads(json_text)
    except Exception as e:
        print("Error parsing statisticsTable JSON:", e)
        exit(1)

    titles = data.get("titles", [])
    rows_data = [item.get("data", []) for item in data.get("items", [])]

    # ============================================================
    # 2. Build a DataFrame, sort by Label, and export to Excel
    # ============================================================
    df = pd.DataFrame(rows_data, columns=titles)
    df.sort_values(by="Label", inplace=True)
    filename = "excelReport.xlsx"
    while (True):
        try:
            df.to_excel(filename, index=False)
            break
        except Exception:
            input("The file excelReport.xlsx is open, close it and press enter to continue")

    print(Fore.GREEN + " The file 'excelReport.xlsx' was created with the statistics table.")

    # ============================================================
    # 3. Create two sheets: "general info" and "metrics"
    # ============================================================
    wb = load_workbook(filename)

    # Rename default sheet to "metrics"
    ws_metrics = wb.active
    ws_metrics.title = "metrics"

    # Add a title row to "metrics"
    ws_metrics.insert_rows(1)
    ws_metrics.merge_cells(start_row=1, start_column=1, end_row=1, end_column=df.shape[1])
    ws_metrics.cell(row=1, column=1, value="Statistics").alignment = Alignment(horizontal="center")
    ws_metrics.cell(row=1, column=1).font = Font(size=16, bold=True)
    statistics_header_row = 2
    statistics_data_end = df.shape[0] + 2
    ws_metrics.sheet_properties.outlinePr.summaryBelow = False
    ws_metrics.freeze_panes = "A3"

    # Create "general info" sheet
    ws_general = wb.create_sheet("general info")
    ws_general.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    ws_general.cell(row=1, column=1, value="General Info").alignment = Alignment(horizontal="center")
    ws_general.cell(row=1, column=1).font = Font(size=16, bold=True)

    ws_general.cell(row=3, column=1, value="Expected tps:").font = Font(size=12, bold=True)
    ws_general.cell(row=3, column=2, value="")  # Optional value
    ws_general.cell(row=4, column=1, value="Complete JMeter report:").font = Font(size=12, bold=True)
    ws_general.cell(row=4, column=2, value=report_url).font = Font(size=12)
    ws_general.cell(row=5, column=1, value="Start Time:").font = Font(size=12, bold=True)
    ws_general.cell(row=5, column=2, value=start_time).font = Font(size=12)
    ws_general.cell(row=6, column=1, value="End Time:").font = Font(size=12, bold=True)
    ws_general.cell(row=6, column=2, value=end_time).font = Font(size=12)

    # Auto-fit columns in the "general info" sheet
    for i, col in enumerate(ws_general.iter_cols(min_row=1, max_row=ws_general.max_row), start=1):
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        ws_general.column_dimensions[get_column_letter(i)].width = max_length + 2

    # ============================================================
    # 4. Insert a blank separator row in "metrics" when the Label prefix changes
    # ============================================================
    separator_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    for i in range(statistics_data_end, statistics_header_row + 1, -1):
        current_label = ws_metrics.cell(row=i, column=1).value
        previous_label = ws_metrics.cell(row=i - 1, column=1).value
        if current_label and previous_label:
            current_prefix = str(current_label).split('.')[0]
            previous_prefix = str(previous_label).split('.')[0]
            if current_prefix != previous_prefix:
                ws_metrics.insert_rows(i)
                rd = ws_metrics.row_dimensions[i]
                rd.outlineLevel = 0
                rd.hidden = False
                rd.collapsed = False
                for col in range(1, df.shape[1] + 1):
                    ws_metrics.cell(row=i, column=col).fill = separator_fill

    stats_end = ws_metrics.max_row

    # ============================================================
    # 5. Group rows in "metrics" by Label prefix
    # ============================================================
    summary_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    current_summary = None
    for row_num in range(statistics_header_row + 1, stats_end + 1):
        label_value = ws_metrics.cell(row=row_num, column=1).value
        if label_value is None:
            current_summary = None
            continue
        if "-" not in str(label_value):  # Parent row
            current_summary = row_num
            for cell in ws_metrics[row_num]:
                cell.fill = summary_fill
        else:
            if current_summary is not None:  # Child row
                ws_metrics.row_dimensions[row_num].outlineLevel = 1
                ws_metrics.row_dimensions[row_num].hidden = True
                ws_metrics.row_dimensions[row_num].collapsed = True


    # ============================================================
    # 6. Insert the Errors table in a new sheet called "errors"
    # ============================================================
    errors_search_string = 'createTable($("#errorsTable")'
    position_errors = js_content.find(errors_search_string)
    if position_errors == -1:
        print("Could not find createTable call for errorsTable.")
        exit(1)

    comma_position_errors = js_content.find(',', position_errors)
    if comma_position_errors == -1:
        print("Could not find the comma after the selector for errorsTable.")
        exit(1)

    json_text_errors = extract_json(js_content, comma_position_errors)
    if not json_text_errors:
        print("Failed to extract the JSON block for errorsTable.")
        exit(1)

    try:
        errors_data = json.loads(json_text_errors)
    except Exception as e:
        print("Error parsing errorsTable JSON:", e)
        exit(1)

    errors_titles = errors_data.get("titles", [])
    errors_rows = [item.get("data", []) for item in errors_data.get("items", [])]
    df_errors = pd.DataFrame(errors_rows, columns=errors_titles)

    # Create new sheet for errors
    ws_errors = wb.create_sheet("errors")

    # Title row
    ws_errors.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df_errors.columns))
    ws_errors.cell(row=1, column=1, value="Errors Table").alignment = Alignment(horizontal="center")
    ws_errors.cell(row=1, column=1).font = Font(size=14, bold=True)

    # Insert the DataFrame into the new sheet
    for r_idx, row in enumerate(dataframe_to_rows(df_errors, index=False, header=True), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws_errors.cell(row=r_idx, column=c_idx, value=value)

    # Apply formatting (header and borders)
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    thin_side = Side(border_style="thin", color="000000")
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # Style headers
    for col in range(1, df_errors.shape[1] + 1):
        cell = ws_errors.cell(row=2, column=col)
        cell.fill = header_fill
        cell.border = border_all

    # Style data cells
    errors_data_end = 2 + df_errors.shape[0]
    for row in ws_errors.iter_rows(min_row=3, max_row=errors_data_end,
                                   min_col=1, max_col=df_errors.shape[1]):
        for cell in row:
            cell.border = border_all
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"

    """errors_search_string = 'createTable($("#errorsTable")'
    position_errors = js_content.find(errors_search_string)
    if position_errors == -1:
        print("Could not find createTable call for errorsTable.")
        exit(1)

    comma_position_errors = js_content.find(',', position_errors)
    if comma_position_errors == -1:
        print("Could not find the comma after the selector for errorsTable.")
        exit(1)

    json_text_errors = extract_json(js_content, comma_position_errors)
    if not json_text_errors:
        print("Failed to extract the JSON block for errorsTable.")
        exit(1)

    try:
        errors_data = json.loads(json_text_errors)
    except Exception as e:
        print("Error parsing errorsTable JSON:", e)
        exit(1)

    errors_titles = errors_data.get("titles", [])
    errors_rows = [item.get("data", []) for item in errors_data.get("items", [])]
    df_errors = pd.DataFrame(errors_rows, columns=errors_titles)

    errors_start_row = stats_end + 2
    ws_metrics.merge_cells(start_row=errors_start_row, start_column=1,
                           end_row=errors_start_row, end_column=len(df_errors.columns))
    ws_metrics.cell(row=errors_start_row, column=1, value="Errors Table").alignment = Alignment(horizontal="center")
    ws_metrics.cell(row=errors_start_row, column=1).font = Font(size=14, bold=True)

    for r_idx, row in enumerate(dataframe_to_rows(df_errors, index=False, header=True),
                                start=errors_start_row + 1):
        for c_idx, value in enumerate(row, start=1):
            ws_metrics.cell(row=r_idx, column=c_idx, value=value)
    """
    # ============================================================
    # 7. Style tables in "metrics"
    # ============================================================
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    thin_side = Side(border_style="thin", color="000000")
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # Header borders/fill for statistics
    for col in range(1, df.shape[1] + 1):
        cell = ws_metrics.cell(row=statistics_header_row, column=col)
        cell.fill = header_fill
        cell.border = border_all

    # Data borders for statistics
    for row in ws_metrics.iter_rows(min_row=statistics_header_row + 1, max_row=stats_end,
                                    min_col=1, max_col=df.shape[1]):
        for cell in row:
            cell.border = border_all
    """
    # Header borders/fill for errors
    errors_header_row = errors_start_row + 1
    for col in range(1, df_errors.shape[1] + 1):
        cell = ws_metrics.cell(row=errors_header_row, column=col)
        cell.fill = header_fill
        cell.border = border_all

    errors_data_end = errors_start_row + df_errors.shape[0] + 1
    for row in ws_metrics.iter_rows(min_row=errors_header_row + 1, max_row=errors_data_end,
                                    min_col=1, max_col=df_errors.shape[1]):
        for cell in row:
            cell.border = border_all
    """
    # ============================================================
    # 7.1. Conditional formatting: highlight Error% > SLA and Avg > SLA
    # ============================================================
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Find Error% column index
    error_col_index = next(
        (col for col in range(1, df.shape[1] + 1)
         if str(ws_metrics.cell(row=statistics_header_row, column=col).value).strip().replace(" ",
                                                                                              "").lower() == "error%"),
        None
    )

    marked_transactions = set()
    if error_col_index:
        for row in range(statistics_header_row + 1, stats_end + 1):
            cell = ws_metrics.cell(row=row, column=error_col_index)
            label_cell = ws_metrics.cell(row=row, column=1)
            if cell.value is not None:
                try:
                    value = float(str(cell.value).replace("%", "").strip())
                    if (value > error_sla and
                            ("GET" in label_cell.value or "POST" in label_cell.value or include_parent_transactions)):
                        cell.fill = red_fill
                        parts = label_cell.value.split('.')
                        marked_transactions.add(f'{parts[0]}.{parts[1]}.')
                except Exception as e:
                    print(f"Conversion error at row {row}, col {error_col_index}: {e}")

    # Propagate red highlight to matching parent rows
    if error_col_index:
        for row in range(statistics_header_row + 1, stats_end + 1):
            label = ws_metrics.cell(row=row, column=1).value
            if label in marked_transactions:
                ws_metrics.cell(row=row, column=error_col_index).fill = red_fill

    # Find Average column index
    avg_col_index = next(
        (col for col in range(1, df.shape[1] + 1)
         if str(ws_metrics.cell(row=statistics_header_row, column=col).value).strip().replace(" ",
                                                                                              "").lower() == "average"),
        None
    )

    avg_red_labels = []
    marked_transactions.clear()
    if avg_col_index:
        for row in range(statistics_header_row + 1, stats_end + 1):
            cell = ws_metrics.cell(row=row, column=avg_col_index)
            label_cell = ws_metrics.cell(row=row, column=1)
            if cell.value is not None:
                try:
                    value = float(str(cell.value).replace("%", "").strip())
                    if (value > avg_time_sla and
                            ("GET" in label_cell.value or "POST" in label_cell.value or include_parent_transactions)):
                        cell.fill = red_fill
                        avg_red_labels.append(label_cell.value)
                        parts = label_cell.value.split('.')
                        marked_transactions.add(f'{parts[0]}.{parts[1]}.')
                except Exception as e:
                    print(f"Conversion error at row {row}, col {avg_col_index}: {e}")

    # Propagate red highlight for Avg column as well
    if avg_col_index:
        for row in range(statistics_header_row + 1, stats_end + 1):
            label = ws_metrics.cell(row=row, column=1).value
            if label in marked_transactions:
                ws_metrics.cell(row=row, column=avg_col_index).fill = red_fill

    # ============================================================
    # 7.2. Format numeric cells with two decimals
    # ============================================================
    for row in ws_metrics.iter_rows(min_row=statistics_header_row + 1, max_row=stats_end,
                                    min_col=1, max_col=df.shape[1]):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"
    """
    for row in ws_metrics.iter_rows(min_row=errors_header_row + 1, max_row=errors_data_end,
                                    min_col=1, max_col=df_errors.shape[1]):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"
    """
    # ============================================================
    # 8. Auto-fit columns on every sheet
    # ============================================================
    for ws in wb.worksheets:
        fit_columns(ws)

    # ============================================================
    # 9. Make "general info" the first sheet
    # ============================================================
    wb._sheets.remove(ws_general)
    wb._sheets.insert(0, ws_general)

    # ============================================================
    # 12. Download graph.js and extract pure JSON from 'data'
    # ============================================================
    graph_url = "https://anlly1796.github.io/jmeter_reports/OrangeHRM/content/js/graph.js"
    graph_response = requests.get(graph_url)
    if graph_response.status_code != 200:
        print("Error retrieving graph.js, status code:", graph_response.status_code)
        exit(1)
    graph_js = graph_response.text

    search_graph = "var responseTimesOverTimeInfos"
    pos_graph = graph_js.find(search_graph)
    if pos_graph == -1:
        print("responseTimesOverTimeInfos not found in graph.js")
        exit(1)
    equal_pos = graph_js.find("=", pos_graph)
    if equal_pos == -1:
        print("'=' not found after responseTimesOverTimeInfos")
        exit(1)
    json_graph_text = extract_json(graph_js, equal_pos)
    if not json_graph_text:
        print("Failed to extract JSON from graph.js")
        exit(1)

    search_data = "data:"
    pos_data = json_graph_text.find(search_data)
    if pos_data == -1:
        print("'data:' not found in extracted JSON from graph.js")
        exit(1)
    start_index_data = pos_data + len(search_data)
    pure_data_text = extract_json(json_graph_text, start_index_data)
    if not pure_data_text:
        print("Failed to extract pure JSON data from graph.js")
        exit(1)

    try:
        graph_data = json.loads(pure_data_text)
    except Exception as e:
        print("Error parsing pure JSON from graph.js:", e)
        exit(1)

    # ============================================================
    # Helper: retrieve response-time array for a given label
    # ============================================================
    def get_response_data_by_label(label: str):
        """
        Return the data array corresponding to the provided label.
        """
        series = graph_data.get("result", {}).get("series", [])
        for s in series:
            if s.get("label") == label:
                return s.get("data", [])
        return None

    # ============================================================
    # 12.1. (Optional) build line charts for every SLA-breaching label
    # ============================================================
    def create_graph_sheet(workbook, data_pairs, label):
        # Sort by epoch timestamp
        data_pairs.sort(key=lambda x: x[0])

        # Convert epoch (ms) to readable HH:MM:SS
        for row in data_pairs:
            epoch_time = row[0] / 1000
            row[0] = datetime.fromtimestamp(epoch_time).strftime("%H:%M:%S")

        ws = workbook.create_sheet(label)
        start_row, start_col = 50, 10  # Hide raw data off-screen

        # Headers
        ws.cell(row=start_row, column=start_col, value="Time")
        ws.cell(row=start_row, column=start_col + 1, value="Value")

        # Data
        for i, row in enumerate(data_pairs, start=1):
            ws.cell(row=start_row + i, column=start_col, value=row[0])
            ws.cell(row=start_row + i, column=start_col + 1, value=row[1])

        # Line chart
        chart = LineChart()
        values = Reference(ws, min_col=start_col + 1, min_row=start_row + 1,
                           max_col=start_col + 1, max_row=start_row + len(data_pairs))
        categories = Reference(ws, min_col=start_col, min_row=start_row + 1,
                               max_row=start_row + len(data_pairs))
        chart.add_data(values, titles_from_data=True)
        chart.set_categories(categories)

        chart.y_axis.majorGridlines = None
        chart.x_axis.majorGridlines = None
        chart.x_axis.title = "Time"
        chart.y_axis.title = "Value"
        chart.x_axis.tickLblPos = "low"
        chart.y_axis.tickLblPos = "low"

        y_values = [row[1] for row in data_pairs]
        chart.y_axis.scaling.min = min(y_values)
        chart.y_axis.scaling.max = max(y_values)

        chart.legend = None
        series = chart.series[0]
        series.graphicalProperties.line.solidFill = "0000FF"
        series.graphicalProperties.line.width = 20000
        series.marker.symbol = "circle"
        series.marker.size = 7
        series.marker.graphicalProperties.solidFill = "0000FF"

        ws.add_chart(chart, "D2")

    # Uncomment if you want to generate charts:
    # for label in avg_red_labels:
    #     data_pairs = get_response_data_by_label(label) or []
    #     if data_pairs:
    #         create_graph_sheet(wb, data_pairs, label)

    # ============================================================
    # 13. Save the Excel file
    # ============================================================
    wb.save(filename)

